from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

def write_excel(df,path,audits=None):
    df.to_excel(path,index=False); wb=load_workbook(path); ws=wb.active; ws.freeze_panes="A2"; ws.auto_filter.ref=ws.dimensions
    for cell in ws[1]:cell.font=Font(bold=True)
    for col in range(1,ws.max_column+1):
        letter=get_column_letter(col); header=str(ws.cell(1,col).value or ""); width=min(60,max(12,len(header)+2))
        for r in range(2,ws.max_row+1):
            v=ws.cell(r,col).value
            if header.lower().endswith("price") or header.lower().endswith("shipping") or header.lower().endswith("total"):ws.cell(r,col).number_format='0.00'
            if "link" in header.lower() and isinstance(v,str) and v.startswith("http"):ws.cell(r,col).hyperlink=v; ws.cell(r,col).style="Hyperlink"; width=28
            width=max(width,min(60,len(str(v or ""))+2))
        ws.column_dimensions[letter].width=width
    if audits:
        aw=wb.create_sheet("Audit"); rows=[]
        for key,market_audit in audits:
            for x in market_audit:rows.append({"row_key":key,**x})
        if rows:
            headers=list(rows[0]); aw.append(headers)
            for x in rows:aw.append([x.get(h) for h in headers])
            aw.freeze_panes="A2"; aw.auto_filter.ref=aw.dimensions
    wb.save(path)
